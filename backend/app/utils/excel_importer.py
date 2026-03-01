"""Import portfolio data from the Excel tracker spreadsheet."""

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from sqlmodel import Session, select

from app.models.fund import Fund, FundCategory
from app.models.institution import Institution
from app.models.manual_asset import ManualAsset
from app.models.target_allocation import TargetAllocation
from app.models.transaction import Transaction


def _safe_decimal(val) -> Decimal | None:
    """Convert a cell value to Decimal, returning None for non-numeric values."""
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def import_excel(
    file_path: str | Path,
    portfolio_id: int,
    session: Session,
) -> dict:
    """Import transactions, funds, and assets from the Excel portfolio tracker.

    Returns summary of imported records.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    summary = {
        "institutions": 0,
        "categories": 0,
        "funds": 0,
        "transactions": 0,
        "manual_assets": 0,
        "target_allocations": 0,
    }

    # --- Import institutions ---
    institutions_map: dict[str, int] = {}
    ws_data = wb["Data"]
    inst_names = set()
    for row in ws_data.iter_rows(min_row=2, max_col=13, values_only=False):
        vals = [c.value for c in row]
        if vals[5]:
            inst_names.add(str(vals[5]).strip())

    for name in inst_names:
        existing = session.exec(
            select(Institution).where(Institution.name == name)
        ).first()
        if existing:
            institutions_map[name] = existing.id
        else:
            inst = Institution(
                name=name,
                institution_type="Bank" if "bank" in name.lower() else "AMC",
            )
            session.add(inst)
            session.commit()
            session.refresh(inst)
            institutions_map[name] = inst.id
            summary["institutions"] += 1

    # --- Import categories ---
    categories_map: dict[str, int] = {}
    cat_names = set()
    for row in ws_data.iter_rows(min_row=2, max_col=13, values_only=False):
        vals = [c.value for c in row]
        if vals[2]:
            cat_names.add(str(vals[2]).strip())

    for name in cat_names:
        existing = session.exec(
            select(FundCategory).where(FundCategory.name == name)
        ).first()
        if existing:
            categories_map[name] = existing.id
        else:
            cat = FundCategory(name=name)
            session.add(cat)
            session.commit()
            session.refresh(cat)
            categories_map[name] = cat.id
            summary["categories"] += 1

    # --- Import funds ---
    funds_map: dict[str, int] = {}
    fund_rows = set()
    for row in ws_data.iter_rows(min_row=2, max_col=13, values_only=False):
        vals = [c.value for c in row]
        if vals[1]:
            scheme_key = str(vals[1]).strip()
            fund_type = str(vals[4]).strip() if vals[4] else "Unknown"
            category = str(vals[2]).strip() if vals[2] else None
            institution = str(vals[5]).strip() if vals[5] else None
            fund_rows.add((scheme_key, fund_type, category, institution))

    for scheme_key, fund_type, category, institution in fund_rows:
        existing = session.exec(
            select(Fund).where(Fund.scheme_key == scheme_key)
        ).first()
        if existing:
            funds_map[scheme_key] = existing.id
        else:
            # Extract fund name from scheme_key: "Category | Fund Name"
            parts = scheme_key.split("|")
            name = parts[-1].strip() if len(parts) > 1 else scheme_key
            fund = Fund(
                scheme_key=scheme_key,
                name=name,
                fund_type=fund_type,
                category_id=categories_map.get(category) if category else None,
                institution_id=institutions_map.get(institution) if institution else None,
                is_shariah_compliant=True,
            )
            session.add(fund)
            session.commit()
            session.refresh(fund)
            funds_map[scheme_key] = fund.id
            summary["funds"] += 1

    # --- Import transactions ---
    for row in ws_data.iter_rows(min_row=2, max_col=13, values_only=False):
        vals = [c.value for c in row]
        if not vals[0]:
            continue

        txn_date = vals[0]
        if isinstance(txn_date, datetime):
            txn_date = txn_date.date()
        elif isinstance(txn_date, str):
            txn_date = date.fromisoformat(txn_date)

        scheme_key = str(vals[1]).strip()
        fund_id = funds_map.get(scheme_key)
        if not fund_id:
            continue

        txn = Transaction(
            portfolio_id=portfolio_id,
            fund_id=fund_id,
            date=txn_date,
            transaction_type=str(vals[6]).strip() if vals[6] else "Deposit",
            units=Decimal(str(vals[7])) if vals[7] else Decimal("0"),
            price_per_unit=Decimal(str(vals[8])) if vals[8] else Decimal("0"),
            amount=Decimal(str(vals[9])) if vals[9] else Decimal("0"),
            signed_amount=Decimal(str(vals[11])) if vals[11] else Decimal("0"),
            xirr_cashflow=Decimal(str(vals[12])) if vals[12] else Decimal("0"),
        )
        session.add(txn)
        summary["transactions"] += 1

    session.commit()

    # --- Recalculate holdings for all imported funds ---
    from app.services.transaction_service import TransactionService
    txn_svc = TransactionService(session)
    for fund_id in set(funds_map.values()):
        if fund_id is not None:
            txn_svc._recalculate_holding(portfolio_id, fund_id)
    summary["holdings_recalculated"] = len([fid for fid in funds_map.values() if fid])

    # --- Import manual assets ---
    if "Manual_Assets" in wb.sheetnames:
        ws_manual = wb["Manual_Assets"]
        for row in ws_manual.iter_rows(min_row=2, max_col=3, values_only=False):
            vals = [c.value for c in row]
            if not vals[0]:
                continue
            asset = ManualAsset(
                portfolio_id=portfolio_id,
                description=str(vals[0]).strip(),
                asset_type="Savings Account",
                current_balance=Decimal(str(vals[1])) if vals[1] else Decimal("0"),
            )
            session.add(asset)
            summary["manual_assets"] += 1
        session.commit()

    # --- Import target allocations ---
    if "Target_Allocation" in wb.sheetnames:
        ws_target = wb["Target_Allocation"]
        for row in ws_target.iter_rows(min_row=2, max_row=8, max_col=2, values_only=False):
            vals = [c.value for c in row]
            if not vals[0] or not vals[1]:
                continue
            ta = TargetAllocation(
                portfolio_id=portfolio_id,
                asset_type=str(vals[0]).strip(),
                target_pct=Decimal(str(vals[1])),
            )
            session.add(ta)
            summary["target_allocations"] += 1
        session.commit()

    # --- Import MUFAP fund data (NAVs + returns) ---
    if "MUFAP_Raw" in wb.sheetnames:
        ws_mufap = wb["MUFAP_Raw"]
        mufap_count = 0
        for row in ws_mufap.iter_rows(min_row=2, max_col=19, values_only=False):
            vals = [c.value for c in row]
            if not vals[0]:
                continue
            scheme_key = str(vals[0]).strip()
            existing = session.exec(
                select(Fund).where(Fund.scheme_key == scheme_key)
            ).first()
            if existing:
                # Update NAV and returns
                nav = _safe_decimal(vals[7])
                if nav is not None:
                    existing.current_nav = nav
                if vals[6]:
                    nav_date = vals[6]
                    if isinstance(nav_date, datetime):
                        existing.nav_updated_at = nav_date
                for attr, idx in [
                    ("return_ytd", 8), ("return_mtd", 9),
                    ("return_30d", 12), ("return_90d", 13),
                    ("return_180d", 14), ("return_365d", 16),
                    ("return_2y", 17), ("return_3y", 18),
                ]:
                    val = _safe_decimal(vals[idx])
                    if val is not None:
                        setattr(existing, attr, val)
                session.add(existing)
            else:
                parts = scheme_key.split("|")
                name = parts[-1].strip() if len(parts) > 1 else scheme_key
                sector = str(vals[1]).strip() if vals[1] else None
                category_name = str(vals[2]).strip() if vals[2] else None
                fund_name_display = str(vals[3]).strip() if vals[3] else name
                rating = str(vals[4]).strip() if vals[4] else None

                cat_id = categories_map.get(category_name) if category_name else None
                if category_name and not cat_id:
                    # Check DB first — may exist from a previous partial import
                    existing_cat = session.exec(
                        select(FundCategory).where(FundCategory.name == category_name)
                    ).first()
                    if existing_cat:
                        categories_map[category_name] = existing_cat.id
                        cat_id = existing_cat.id
                    else:
                        cat = FundCategory(name=category_name, sector=sector)
                        session.add(cat)
                        session.commit()
                        session.refresh(cat)
                        categories_map[category_name] = cat.id
                        cat_id = cat.id

                # Determine fund_type from category
                fund_type = "Unknown"
                if category_name:
                    cl = category_name.lower()
                    if "equity" in cl:
                        fund_type = "Equity Fund"
                    elif "money market" in cl:
                        fund_type = "Money Market Fund"
                    elif "income" in cl or "debt" in cl or "sukuk" in cl:
                        fund_type = "Debt Fund"
                    elif "pension" in cl and "equity" in cl:
                        fund_type = "Pension Fund - Equity"
                    elif "pension" in cl and "debt" in cl:
                        fund_type = "Pension Fund - Debt"
                    elif "gold" in cl or "commodity" in cl:
                        fund_type = "Gold Fund"
                    elif "balanced" in cl or "asset allocation" in cl:
                        fund_type = "Balanced Fund"
                    elif "shariah" in cl:
                        fund_type = "Shariah Fund"

                is_shariah = "shariah" in scheme_key.lower() or "islamic" in scheme_key.lower()

                fund = Fund(
                    scheme_key=scheme_key,
                    name=fund_name_display,
                    fund_type=fund_type,
                    category_id=cat_id,
                    rating=rating,
                    is_shariah_compliant=is_shariah,
                    current_nav=_safe_decimal(vals[7]),
                    nav_updated_at=vals[6] if isinstance(vals[6], datetime) else None,
                    return_ytd=_safe_decimal(vals[8]),
                    return_mtd=_safe_decimal(vals[9]),
                    return_30d=_safe_decimal(vals[12]),
                    return_90d=_safe_decimal(vals[13]),
                    return_180d=_safe_decimal(vals[14]),
                    return_365d=_safe_decimal(vals[16]),
                    return_2y=_safe_decimal(vals[17]),
                    return_3y=_safe_decimal(vals[18]),
                )
                session.add(fund)
                funds_map[scheme_key] = None  # ID assigned after commit
                mufap_count += 1

        session.commit()
        summary["mufap_funds_imported"] = mufap_count

    # --- Import NAV_Master data ---
    if "NAV_Master" in wb.sheetnames:
        ws_nav = wb["NAV_Master"]
        for row in ws_nav.iter_rows(min_row=2, max_col=4, values_only=False):
            vals = [c.value for c in row]
            if not vals[0]:
                continue
            scheme_key = str(vals[0]).strip()
            fund = session.exec(
                select(Fund).where(Fund.scheme_key == scheme_key)
            ).first()
            if fund and vals[2] is not None:
                fund.current_nav = Decimal(str(vals[2]))
                if vals[3] and isinstance(vals[3], datetime):
                    fund.nav_updated_at = vals[3]
                session.add(fund)
        session.commit()

    return summary
